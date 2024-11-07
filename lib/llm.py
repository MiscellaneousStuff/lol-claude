# Standard library imports
import base64
import json
import os
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path

# Third-party imports
import aiofiles
import pillow_heif
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance

# Register HEIF opener
pillow_heif.register_heif_opener()

# DOTENV
from dotenv import load_dotenv
load_dotenv()

# LLMs
from openai import OpenAI
from anthropic import Anthropic
gpt_client = OpenAI()
claude_client = Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

# PDF to Image
import pdf2image

# Typing
from typing import List

# Async
import aiohttp
import asyncio

# Register the HEIF opener with Pillow
pillow_heif.register_heif_opener()

# AWS Bedrock (Claude)
from anthropic import AnthropicBedrock
bedrock_client = AnthropicBedrock(
    aws_access_key=os.getenv("AWS_ACCESS_KEY_ID"),
    aws_secret_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
    aws_region=os.getenv("AWS_REGION")
)

from pdf2image import convert_from_path

import xml.etree.ElementTree as ET

async def claude(txt: str, path: str = "", temperature: float = 0.7):
    """
    Sends a request to the Claude AI model with text and optional image input.

    This function prepares the content for a request to the Anthropic API, including
    text and optional image data. It handles both PDF and image file inputs,
    converting PDFs to images when necessary.

    Args:
        txt (str): The text prompt to send to Claude.
        path (str, optional): Path to an image or PDF file to include in the request. Defaults to "".
        temperature (float, optional): The sampling temperature for the AI model. Defaults to 0.7.

    Returns:
        str: The response from the Claude AI model.

    Raises:
        HTTPException: If there's an error with the Anthropic API request or response.

    Note:
        This function requires the ANTHROPIC_API_KEY environment variable to be set.
    """
    path = str(path)
    content = [
        {
            "type": "text",
            "text": txt
        }
    ]
    if path:
        if path.endswith(".pdf"):
            # Convert PDF to images
            image_paths = await pdf_to_images(path)
        else:
            # For single image files
            image_paths = [path]
        for img_path in image_paths:
            # Process the image
            processed_img_path = await process_image(img_path, scaling=1)
            base64_image = await encode_image(processed_img_path)
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/jpeg",
                    "data": base64_image
                }
            })

    async with aiohttp.ClientSession() as session:
        try:
            async with session.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "Content-Type": "application/json",
                    "X-API-Key": os.environ.get("ANTHROPIC_API_KEY"),
                    "anthropic-version": "2023-06-01"  # Add the required header
                },
                json={
                    "max_tokens": 4096,
                    "messages": [{"role": "user", "content": content}],
                    "model": "claude-3-5-sonnet-20240620",
                    "temperature": temperature,
                },
            ) as response:
                result = await response.json()
                if response.status != 200:
                    error_message = result.get('error', {}).get('message', 'Unknown error occurred')
                    raise HTTPException(status_code=response.status, detail=f"Anthropic API error: {error_message}")
                
                if 'content' not in result or not result['content']:
                    raise HTTPException(status_code=500, detail="Unexpected response format from Anthropic API")
                
                return result["content"][0]["text"]
        except aiohttp.ClientError as e:
            raise HTTPException(status_code=500, detail=f"Error communicating with Anthropic API: {str(e)}")

async def bedrock_claude(txt: str, path: str = "", temperature: float = 0.7):
    path = str(path)
    content = [
        {
            "type": "text",
            "text": txt
        }
    ]
    if path:
        if path.lower().endswith(".pdf"):
            # Convert PDF to images
            image_paths = await pdf_to_images(path)
        else:
            # For single image files
            image_paths = [path]
        for img_path in image_paths:
            # Process the image
            processed_img_path = await process_image(img_path, scaling=1)
            base64_image = await encode_image(processed_img_path)
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/jpeg",
                    "data": base64_image
                }
            })

    try:
        response = await asyncio.to_thread(
            bedrock_client.messages.create,
            max_tokens=4096,
            messages=[{"role": "user", "content": content}],
            model="anthropic.claude-3-sonnet-20240229-v1:0",
            temperature=temperature,
        )
        
        if not response.content:
            raise HTTPException(status_code=500, detail="Unexpected response format from Bedrock API")
        
        return response.content[0].text
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error communicating with Bedrock API: {str(e)}")
        
async def pdf_to_images(pdf_path: str) -> List[str]:
    """
    Converts a PDF file to a list of image paths.

    This asynchronous function takes a PDF file path as input and converts each page
    of the PDF into a separate JPEG image. It uses the pdf2image library to perform
    the conversion in a separate thread to avoid blocking the event loop.

    Args:
        pdf_path (str): The file path of the PDF to be converted.

    Returns:
        List[str]: A list of file paths for the generated JPEG images, one for each
                   page of the PDF.

    Raises:
        Any exceptions raised by pdf2image.convert_from_path or image processing
        operations will be propagated.

    Note:
        This function creates temporary JPEG files in the same directory as the
        input PDF, named with the pattern "{pdf_path}_page_{page_number}.jpg".
        These files are not automatically deleted and should be managed by the caller.
    """

    # This function is CPU-bound, so we'll run it in a separate thread
    loop = asyncio.get_event_loop()
    images = await loop.run_in_executor(None, convert_from_path, pdf_path)
    
    image_paths = []
    for i, image in enumerate(images):
        image_path = f"{pdf_path}_page_{i+1}.jpg"
        # Convert to RGB before saving
        rgb_image = await loop.run_in_executor(None, lambda: image.convert('RGB'))
        await loop.run_in_executor(None, rgb_image.save, image_path, "JPEG")
        image_paths.append(image_path)
    
    return image_paths

async def process_image(image_path: str, scaling: float = 1, max_size: int = 2000) -> str:
    """
    Process an image file by resizing it if necessary.

    This asynchronous function takes an image file path as input and processes the image
    by resizing it based on the given scaling factor and maximum size constraints.

    Args:
        image_path (str): The file path of the input image.
        scaling (float, optional): The scaling factor to apply to the image. Defaults to 1.
        max_size (int, optional): The maximum allowed dimension (width or height) of the image. Defaults to 2000.

    Returns:
        str: The file path of the processed image.

    The function performs the following steps:
    1. Opens the image file.
    2. Calculates new dimensions based on the max_size and scaling parameters.
    3. Resizes the image if necessary, maintaining the aspect ratio.
    4. Saves the processed image as a JPEG file.

    Note:
    - The function uses asyncio to run CPU-bound operations in a separate thread.
    - The processed image is saved with a "_processed.jpg" suffix added to the original filename.
    """
    loop = asyncio.get_event_loop()
    with await loop.run_in_executor(None, Image.open, image_path) as img:
        # Convert RGBA to RGB if necessary
        if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
            # Create a white background
            background = Image.new('RGB', img.size, (255, 255, 255))
            # Paste the image on the background using alpha channel as mask
            if img.mode == 'RGBA':
                background.paste(img, mask=img.split()[3])
            else:
                background.paste(img)
            img = background
            
        width, height = img.size
        
        # Calculate new dimensions while maintaining aspect ratio
        if width > max_size or height > max_size:
            ratio = min(max_size / width, max_size / height)
            new_size = (int(width * ratio), int(height * ratio))
        elif scaling != 1:
            new_size = (int(width * scaling), int(height * scaling))
        else:
            new_size = (width, height)
        
        # Resize only if necessary
        if new_size != (width, height):
            img = await loop.run_in_executor(None, img.resize, new_size, Image.LANCZOS)
        
        output_path = f"{image_path}_processed.jpg"
        await loop.run_in_executor(None, img.save, output_path, "JPEG")
    
    return output_path

async def encode_image(image_path: str) -> str:
    """
    Encode an image file to base64 string.

    This asynchronous function takes an image file path as input and encodes
    the image data to a base64 string.

    Args:
        image_path (str): The file path of the input image.

    Returns:
        str: The base64 encoded string representation of the image.

    The function performs the following steps:
    1. Opens the image file asynchronously.
    2. Reads the binary data of the image.
    3. Encodes the binary data to a base64 string.
    4. Decodes the base64 bytes to a UTF-8 string.

    Note:
    - This function uses aiofiles for asynchronous file I/O operations.
    - The returned string is ready to be used in data URIs or for transmission.
    """
    async with aiofiles.open(image_path, "rb") as image_file:
        image_data = await image_file.read()
    
    return base64.b64encode(image_data).decode('utf-8')

def gpt(txt, path="", temperature=0.7):
    """
    Send a text prompt to the GPT model and optionally include image data.

    This function sends a text prompt to the GPT model and can include image data
    if a file path is provided. It supports both PDF and image files.

    Args:
        txt (str): The text prompt to send to the model.
        path (str, optional): The file path of an image or PDF to include. Defaults to "".
        temperature (float, optional): The sampling temperature for the model. Defaults to 0.7.

    Returns:
        str: The response content from the GPT model.

    The function performs the following steps:
    1. Prepares the content list with the text prompt.
    2. If a file path is provided:
       - For PDFs, converts pages to images.
       - For single images, uses the path directly.
    3. Processes and encodes each image.
    4. Sends the prepared content to the GPT model.
    5. Returns the model's response.

    Note:
    - The function uses external functions like pdf_to_images, process_image, and encode_image.
    - It assumes the existence of a gpt_client object for API communication.
    """
    model = "gpt-4o"
    content = [{"type": "text", "text": txt}]
    path = str(path)
    if path:
        if path.endswith(".pdf"):
            # Convert PDF to images
            image_paths = pdf_to_images(path)
        else:
            # For single image files
            image_paths = [path]
        for img_path in image_paths:
            # Process the image
            processed_img_path = process_image(img_path, scaling=1.0)
            base64_image = encode_image(processed_img_path)
            content.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{base64_image}"
                }
            })
    response = gpt_client.chat.completions.create(
        messages=[ {"role": "user", "content": content} ],
        model=model,
        temperature=temperature,
        max_tokens=4_096
    )
    msg = response.choices[0].message.content
    return msg

def write_json_to_xlsx(json_data: object, output_file: str):
    """
    Write JSON data to an Excel (.xlsx) file.

    This function takes a JSON object and writes its contents to an Excel file.
    It creates a new workbook, writes headers based on the JSON keys, and then
    populates the rows with the corresponding values.

    Args:
        json_data (object): The JSON data to be written to the Excel file.
            Expected to be a list of dictionaries or a list of lists of dictionaries.
        output_file (str): The path and filename for the output Excel file.

    Returns:
        None

    The function performs the following steps:
    1. Creates a new workbook and selects the active sheet.
    2. Flattens the JSON data if it's a list of lists.
    3. Writes headers based on the keys of the first JSON object.
    4. Writes data rows for each item in the JSON data.
    5. Adjusts column widths for better readability.
    6. Saves the workbook to the specified output file.

    Note:
    - If the JSON data is empty, a message is printed and no file is created.
    - The function uses the openpyxl library for Excel file operations.
    - Headers are formatted with bold font and center alignment.
    - All columns are set to a width of 20 characters.
    """
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Flatten the list of lists if necessary
    if isinstance(json_data[0], list):
        json_data = [item for sublist in json_data for item in sublist]

    # Check if json_data is empty
    if not json_data:
        print("No data to write to Excel.")
        return

    # Write headers
    headers = list(json_data[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    for row, item in enumerate(json_data, start=2):
        for col, key in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=item.get(key, ''))

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook
    wb.save(output_file)
    print(f"Data written to {output_file}")